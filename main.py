from collections import defaultdict
from datetime import datetime
import os
import threading
import time

import cv2
import numpy as np
import requests
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
import sqlite3
from fastapi.security import OAuth2PasswordBearer
from fastapi import Depends, HTTPException, status
from auth import (
    get_password_hash, 
    verify_password, 
    create_access_token, 
    create_refresh_token, 
    decode_token
)
import random
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from google.oauth2 import id_token
from google.auth.transport import requests as google_requests

# =========================
# GOOGLE OAUTH CONFIG
# =========================
GOOGLE_CLIENT_ID = "158844941712-ur7qfnon9j69tg1vkobfdbip5a4ujm0e.apps.googleusercontent.com"

app = FastAPI()

# =========================
# CONFIGURACION GENERAL
# =========================
# Opciones: "IP" o "LOCAL"
CAMERA_TYPE = "LOCAL" 
# URL para DroidCam (si CAMERA_TYPE es "IP")
DROIDCAM_URL = "http://192.168.1.75:4747/video"
# Índice de cámara (si CAMERA_TYPE es "LOCAL")
# 0 suele ser la cámara de la laptop, 1 o más suelen ser cámaras USB
CAMERA_INDEX = 0

LUGAR = "Libreria Leon"
SAVE_COOLDOWN_SECONDS = 5

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "registro_rostros.xlsx")
MODELS_DIR = os.path.join(BASE_DIR, "models")

FACE_PROTO = os.path.join(MODELS_DIR, "opencv_face_detector.pbtxt")
FACE_MODEL = os.path.join(MODELS_DIR, "opencv_face_detector_uint8.pb")

GENDER_PROTO = os.path.join(MODELS_DIR, "gender_deploy.prototxt")
GENDER_MODEL = os.path.join(MODELS_DIR, "gender_net.caffemodel")

GENDER_LIST = ["Hombre", "Mujer"]
MODEL_MEAN_VALUES = (78.4263377603, 87.7689143744, 114.895847746)

# =========================
# CORS PARA ANGULAR
# =========================
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:4200",
        "http://127.0.0.1:4200",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================
# VARIABLES GLOBALES
# =========================
latest_frame = None
frame_lock = threading.Lock()
running = True

last_faces = []
last_detection_state = False
last_save_time = 0.0
last_frame_received_time = 0.0

face_net = None
gender_net = None

stable_faces = []
face_memory = {}
FACE_STABILITY_FRAMES = 5
FACE_MATCH_DISTANCE = 50
FRAME_TIMEOUT_SECONDS = 3.0

net_lock = threading.Lock()
excel_lock = threading.Lock()
state_lock = threading.Lock()

# =========================
# CONFIGURACION AUTH
# =========================
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="auth/login")
DB_FILE = os.path.join(BASE_DIR, "users.db")
# =========================
# CONFIGURACION EMAIL (SMTP)
# =========================
SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", 587))
SMTP_USER = os.getenv("SMTP_USER", "alexander2004deleon@gmail.com")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "rgpm vvez pczr nmbz")

def send_2fa_email(target_email: str, code: str):
    """Envía el código 2FA por correo electrónico."""
    try:
        msg = MIMEMultipart()
        msg['From'] = SMTP_USER
        msg['To'] = target_email
        msg['Subject'] = f"Tu código de acceso - {LUGAR}"

        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; text-align: center; padding: 20px;">
                <h2 style="color: #2563eb;">Verificación de Seguridad</h2>
                <p>Has intentado iniciar sesión en <strong>{LUGAR}</strong>.</p>
                <p>Tu código de verificación de 6 dígitos es:</p>
                <div style="background: #f3f4f6; padding: 20px; border-radius: 10px; display: inline-block; font-size: 32px; font-weight: bold; letter-spacing: 5px; color: #111827; margin: 20px 0;">
                    {code}
                </div>
                <p style="color: #6b7280; font-size: 14px;">Este código expirará en 5 minutos.</p>
                <p style="color: #6b7280; font-size: 12px; margin-top: 40px;">Si no solicitaste este código, puedes ignorar este correo.</p>
            </body>
        </html>
        """
        msg.attach(MIMEText(body, 'html'))

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)
        
        print(f"Correo enviado exitosamente a {target_email}")
        return True
    except Exception as e:
        print(f"Error enviando correo: {e}")
        return False

def send_reset_password_email(target_email: str, code: str):
    """Envía el código de recuperación de contraseña por correo electrónico."""
    try:
        msg = MIMEMultipart()
        msg['From'] = SMTP_USER
        msg['To'] = target_email
        msg['Subject'] = f"Recuperar contraseña - {LUGAR}"

        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; text-align: center; padding: 20px;">
                <h2 style="color: #ef4444;">Recuperación de Contraseña</h2>
                <p>Has solicitado restablecer tu contraseña en <strong>{LUGAR}</strong>.</p>
                <p>Tu código de recuperación es:</p>
                <div style="background: #fef2f2; padding: 20px; border-radius: 10px; display: inline-block; font-size: 36px; font-weight: bold; letter-spacing: 8px; color: #b91c1c; margin: 20px 0;">
                    {code}
                </div>
                <p style="color: #6b7280; font-size: 14px;">Este código expirará en <strong>1 minuto</strong>.</p>
                <p style="color: #6b7280; font-size: 12px; margin-top: 40px;">Si no solicitaste este cambio, puedes ignorar este correo con seguridad.</p>
            </body>
        </html>
        """
        msg.attach(MIMEText(body, 'html'))

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)
        
        return True
    except Exception as e:
        print(f"Error enviando correo de recuperación: {e}")
        return False

pending_2fa = {} # {username: {"code": "123456", "expires": timestamp}}
pending_reset = {} # {email: {"code": "123456", "expires": timestamp}}

@app.post("/auth/forgot-password")
def forgot_password(data: dict):
    email = data.get("email")
    if not email:
        raise HTTPException(status_code=400, detail="El correo es requerido")

    # Verificar si el usuario existe
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM users WHERE email = ?", (email,))
    user = cursor.fetchone()
    conn.close()

    if not user:
        # Por seguridad, no revelamos si el correo existe o no, pero aquí 
        # para facilidad del usuario devolveremos un error claro
        raise HTTPException(status_code=404, detail="No existe un usuario con ese correo")

    # Generar código de recuperación
    code = f"{random.randint(10000, 99999)}"
    pending_reset[email] = {
        "code": code,
        "expires": time.time() + 60  # 1 minuto
    }

    # Enviar correo
    sent = send_reset_password_email(email, code)
    if not sent:
        raise HTTPException(status_code=500, detail="Error al enviar el correo")

    return {"message": "Código de recuperación enviado"}

@app.post("/auth/reset-password")
def reset_password(data: dict):
    email = data.get("email")
    code = data.get("code")
    new_password = data.get("password")

    if not email or not code or not new_password:
        raise HTTPException(status_code=400, detail="Faltan datos")

    if email not in pending_reset:
        raise HTTPException(status_code=400, detail="No se ha solicitado recuperación para este correo")

    reset_info = pending_reset[email]
    if time.time() > reset_info["expires"]:
        del pending_reset[email]
        raise HTTPException(status_code=400, detail="El código ha expirado")

    if reset_info["code"] != code:
        raise HTTPException(status_code=400, detail="El código es incorrecto")

    # Actualizar contraseña
    hashed_pw = get_password_hash(new_password)
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET password = ? WHERE email = ?", (hashed_pw, email))
    conn.commit()
    conn.close()

    # Limpiar código usado
    del pending_reset[email]

    return {"message": "Contraseña actualizada correctamente"}

def init_db():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

def get_current_user(token: str = Depends(oauth2_scheme)):
    payload = decode_token(token)
    if not payload or "sub" not in payload:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token inválido o expirado",
            headers={"WWW-Authenticate": "Bearer"},
        )
    return payload["sub"]


# =========================
# VALIDACION DE MODELOS
# =========================
def validate_model_files():
    required_files = [
        FACE_PROTO,
        FACE_MODEL,
        GENDER_PROTO,
        GENDER_MODEL,
    ]

    missing_files = [file_path for file_path in required_files if not os.path.exists(file_path)]

    if missing_files:
        raise FileNotFoundError(
            "No se encontraron los siguientes archivos del modelo:\n" +
            "\n".join(missing_files)
        )


def load_models():
    global face_net, gender_net

    validate_model_files()

    print("Cargando modelos...")
    print("FACE_PROTO:", FACE_PROTO)
    print("FACE_MODEL:", FACE_MODEL)
    print("GENDER_PROTO:", GENDER_PROTO)
    print("GENDER_MODEL:", GENDER_MODEL)

    face_net = cv2.dnn.readNet(FACE_MODEL, FACE_PROTO)
    gender_net = cv2.dnn.readNet(GENDER_MODEL, GENDER_PROTO)

    print("Modelos cargados correctamente.")


# =========================
# EXCEL
# =========================
def init_excel():
    try:
        with excel_lock:
            if not os.path.exists(EXCEL_FILE):
                wb = Workbook()
                ws = wb.active
                ws.title = "Registros"
                ws.append(["FechaHora", "Lugar", "RostrosDetectados", "Genero"])
                wb.save(EXCEL_FILE)
                wb.close()
                print(f"Archivo Excel creado: {EXCEL_FILE}")
                return

            wb = load_workbook(EXCEL_FILE)
            ws = wb["Registros"]

            headers = [cell.value for cell in ws[1]]

            if "Genero" not in headers:
                ws.cell(row=1, column=len(headers) + 1, value="Genero")
                wb.save(EXCEL_FILE)
                print("Columna 'Genero' agregada al Excel existente.")

            wb.close()
            print(f"Archivo Excel ya existe: {EXCEL_FILE}")

    except Exception as e:
        print("Error creando o actualizando Excel:", e)


def save_detection_to_excel(num_faces: int, generos: list[str]):
    with excel_lock:
        for intento in range(3):
            wb = None
            try:
                now = datetime.now()

                wb = load_workbook(EXCEL_FILE)
                ws = wb["Registros"]

                generos_texto = ", ".join(generos) if generos else "No identificado"

                ws.append([
                    now.strftime("%Y-%m-%d %H:%M:%S"),
                    LUGAR,
                    num_faces,
                    generos_texto
                ])

                wb.save(EXCEL_FILE)

                print(
                    f"Registro guardado: {now.strftime('%Y-%m-%d %H:%M:%S')} | "
                    f"Rostros: {num_faces} | Generos: {generos_texto}"
                )
                return True

            except PermissionError:
                print(f"Intento {intento + 1}: Excel bloqueado. Reintentando...")
                time.sleep(1)

            except Exception as e:
                print("Error guardando en Excel:", e)
                return False

            finally:
                if wb is not None:
                    try:
                        wb.close()
                    except Exception:
                        pass

        print("No se pudo guardar en Excel después de varios intentos.")
        return False


def read_excel_rows():
    if not os.path.exists(EXCEL_FILE):
        return []

    with excel_lock:
        wb = None
        try:
            wb = load_workbook(EXCEL_FILE, data_only=True)
            ws = wb["Registros"]

            headers = [cell.value for cell in ws[1]]
            rows = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                item = dict(zip(headers, row))
                rows.append(item)

            return rows

        finally:
            if wb is not None:
                wb.close()


# =========================
# CAMERA READER (IP & LOCAL)
# =========================
def camera_reader():
    global latest_frame, running, last_frame_received_time

    while running:
        if CAMERA_TYPE == "IP":
            # Lógica para DroidCam (Stream HTTP)
            session = None
            stream = None
            try:
                print(f"Intentando conectar a Cámara IP: {DROIDCAM_URL}")
                session = requests.Session()
                session.headers.update({
                    "Connection": "close",
                    "User-Agent": "Python-DroidCam-Client"
                })
                stream = session.get(DROIDCAM_URL, stream=True, timeout=(5, 5))

                if stream.status_code != 200:
                    print(f"Error conectando a Cámara IP: HTTP {stream.status_code}")
                    with frame_lock:
                        latest_frame = None
                    time.sleep(2)
                    continue

                print("Conectado a Cámara IP correctamente.")
                bytes_data = b""
                last_chunk_time = time.time()

                for chunk in stream.iter_content(chunk_size=4096):
                    if not running or CAMERA_TYPE != "IP":
                        break

                    now = time.time()
                    if now - last_chunk_time > FRAME_TIMEOUT_SECONDS:
                        print("Timeout de frames detectado. Reconectando...")
                        break

                    if not chunk:
                        time.sleep(0.01)
                        continue

                    last_chunk_time = now
                    bytes_data += chunk
                    a = bytes_data.find(b"\xff\xd8")
                    b = bytes_data.find(b"\xff\xd9")

                    if a != -1 and b != -1 and b > a:
                        jpg = bytes_data[a:b + 2]
                        bytes_data = bytes_data[b + 2:]
                        frame = cv2.imdecode(np.frombuffer(jpg, dtype=np.uint8), cv2.IMREAD_COLOR)
                        if frame is not None:
                            with frame_lock:
                                latest_frame = frame
                                last_frame_received_time = time.time()

            except requests.exceptions.ConnectTimeout:
                print(f"⚠️ Timeout: No se pudo conectar a {DROIDCAM_URL}. Verifica que DroidCam esté abierto y en la misma red.")
                with frame_lock:
                    latest_frame = None
                time.sleep(5)
            except requests.exceptions.ConnectionError:
                print(f"❌ Error de conexión: No se pudo establecer vínculo con {DROIDCAM_URL}.")
                with frame_lock:
                    latest_frame = None
                time.sleep(5)
            except Exception as e:
                print(f"Error en stream IP: {e}")
                with frame_lock:
                    latest_frame = None
                time.sleep(2)
            finally:
                if stream: stream.close()
                if session: session.close()

        else:
            # Lógica para Cámara Local (USB o Laptop)
            cap = None
            try:
                print(f"Intentando conectar a Cámara Local (Índice {CAMERA_INDEX})...")
                cap = cv2.VideoCapture(CAMERA_INDEX)
                
                # Opcional: configurar resolución
                cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
                cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

                if not cap.isOpened():
                    print(f"No se pudo abrir la cámara local índice {CAMERA_INDEX}")
                    with frame_lock:
                        latest_frame = None
                    time.sleep(2)
                    continue

                print(f"Cámara Local {CAMERA_INDEX} conectada.")
                
                while running and CAMERA_TYPE == "LOCAL":
                    ret, frame = cap.read()
                    if not ret:
                        print("Error leyendo frame de cámara local.")
                        break
                    
                    with frame_lock:
                        latest_frame = frame
                        last_frame_received_time = time.time()
                    
                    # Pequeña pausa para no saturar el CPU
                    time.sleep(0.01)

            except Exception as e:
                print(f"Error en cámara local: {e}")
                with frame_lock:
                    latest_frame = None
                time.sleep(2)
            finally:
                if cap:
                    cap.release()

        time.sleep(1)

# =========================
# DETECCION DE ROSTROS
# =========================
def detectar_rostros_dnn(frame, conf_threshold=0.85):
    global face_net

    if face_net is None:
        return []

    h, w = frame.shape[:2]

    # Ignorar una parte superior donde suelen salir techo, focos y reflejos
    y_inicio = int(h * 0.10)
    roi = frame[y_inicio:h, 0:w]
    roi_h, roi_w = roi.shape[:2]

    if roi_h <= 0 or roi_w <= 0:
        return []

    blob = cv2.dnn.blobFromImage(
        image=roi,
        scalefactor=1.0,
        size=(300, 300),
        mean=(104, 117, 123),
        swapRB=False,
        crop=False
    )

    with net_lock:
        face_net.setInput(blob)
        detections = face_net.forward()

    boxes = []
    confidences = []

    min_face_size = 30
    max_face_size_ratio = 0.85
    margin_x = int(roi_w * 0.03)
    margin_y = int(roi_h * 0.03)

    for i in range(detections.shape[2]):
        confidence = float(detections[0, 0, i, 2])

        if confidence < conf_threshold:
            continue

        x1 = int(detections[0, 0, i, 3] * roi_w)
        y1 = int(detections[0, 0, i, 4] * roi_h)
        x2 = int(detections[0, 0, i, 5] * roi_w)
        y2 = int(detections[0, 0, i, 6] * roi_h)

        x1 = max(0, x1)
        y1 = max(0, y1)
        x2 = min(roi_w - 1, x2)
        y2 = min(roi_h - 1, y2)

        box_w = x2 - x1
        box_h = y2 - y1

        if box_w < min_face_size or box_h < min_face_size:
            continue

        if box_w > roi_w * max_face_size_ratio or box_h > roi_h * max_face_size_ratio:
            continue

        aspect_ratio = box_w / float(box_h)
        if aspect_ratio < 0.60 or aspect_ratio > 1.50:
            continue

        if x1 < margin_x or y1 < margin_y or x2 > roi_w - margin_x or y2 > roi_h - margin_y:
            continue

        boxes.append([x1, y1, box_w, box_h])
        confidences.append(confidence)

    if not boxes:
        return []

    indices = cv2.dnn.NMSBoxes(boxes, confidences, conf_threshold, 0.35)

    faces = []
    if len(indices) > 0:
        for idx in indices.flatten():
            x, y, bw, bh = boxes[idx]
            conf = confidences[idx]
            faces.append((x, y + y_inicio, bw, bh, conf))

    faces.sort(key=lambda item: item[4], reverse=True)
    faces = faces[:5]

    return faces


def detectar_genero(face_img):
    global gender_net

    if gender_net is None or face_img is None or face_img.size == 0:
        return "No identificado", 0.0

    try:
        face_img = cv2.resize(face_img, (227, 227))

        blob = cv2.dnn.blobFromImage(
            image=face_img,
            scalefactor=1.0,
            size=(227, 227),
            mean=MODEL_MEAN_VALUES,
            swapRB=False,
            crop=False
        )

        with net_lock:
            gender_net.setInput(blob)
            preds = gender_net.forward()

        gender_index = preds[0].argmax()
        gender = GENDER_LIST[gender_index]
        confidence = float(preds[0][gender_index])

        return gender, confidence

    except Exception as e:
        print("Error detectando genero:", e)
        return "No identificado", 0.0


def estabilizar_rostros(faces_con_genero):
    global face_memory, stable_faces

    nuevas_memorias = {}
    rostros_estables = []

    for (x, y, w, h, face_conf, genero, gender_conf) in faces_con_genero:
        cx = x + w // 2
        cy = y + h // 2

        mejor_key = None
        mejor_dist = None

        for key in face_memory.keys():
            px, py = key
            dist = ((cx - px) ** 2 + (cy - py) ** 2) ** 0.5

            if dist <= FACE_MATCH_DISTANCE and (mejor_dist is None or dist < mejor_dist):
                mejor_dist = dist
                mejor_key = key

        if mejor_key is not None:
            count = face_memory[mejor_key]["count"] + 1
        else:
            count = 1

        nuevas_memorias[(cx, cy)] = {
            "count": count,
            "data": (x, y, w, h, face_conf, genero, gender_conf)
        }

        if count >= FACE_STABILITY_FRAMES:
            rostros_estables.append((x, y, w, h, face_conf, genero, gender_conf))

    face_memory = nuevas_memorias
    stable_faces = rostros_estables


def detect_faces_and_log(frame):
    global last_faces, last_detection_state, last_save_time

    faces = detectar_rostros_dnn(frame)
    faces_con_genero = []

    for (x, y, w, h, face_conf) in faces:
        pad_x = int(w * 0.12)
        pad_y = int(h * 0.12)

        x1 = max(0, x - pad_x)
        y1 = max(0, y - pad_y)
        x2 = min(frame.shape[1], x + w + pad_x)
        y2 = min(frame.shape[0], y + h + pad_y)

        face_crop = frame[y1:y2, x1:x2]

        if face_crop is not None and face_crop.size > 0:
            genero, gender_conf = detectar_genero(face_crop)
        else:
            genero, gender_conf = "No identificado", 0.0

        if gender_conf < 0.70:
            genero = "No identificado"

        faces_con_genero.append((x, y, w, h, face_conf, genero, gender_conf))

    estabilizar_rostros(faces_con_genero)

    has_face = len(stable_faces) > 0
    current_time = time.time()

    with state_lock:
        last_faces = list(stable_faces)

        if has_face:
            if (not last_detection_state) or (current_time - last_save_time >= SAVE_COOLDOWN_SECONDS):
                generos_detectados = [item[5] for item in stable_faces]
                if save_detection_to_excel(len(stable_faces), generos_detectados):
                    last_save_time = current_time

        last_detection_state = has_face


def detection_worker():
    global running

    while running:
        try:
            with frame_lock:
                frame = None if latest_frame is None else latest_frame.copy()

            if frame is not None:
                detect_faces_and_log(frame)

        except Exception as e:
            print("Error en detection_worker:", e)

        time.sleep(0.25)


# =========================
# STREAM DE VIDEO
# =========================
def generate_frames():
    while True:
        try:
            with frame_lock:
                frame_vencido = (
                    last_frame_received_time > 0 and
                    (time.time() - last_frame_received_time) > FRAME_TIMEOUT_SECONDS
                )

                if latest_frame is None or frame_vencido:
                    frame = None
                else:
                    frame = latest_frame.copy()

            if frame is None:
                standby = np.zeros((480, 640, 3), dtype=np.uint8)
                cv2.putText(
                    standby,
                    "Esperando camara...",
                    (170, 240),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    1,
                    (255, 255, 255),
                    2
                )

                ret, buffer = cv2.imencode(
                    ".jpg",
                    standby,
                    [int(cv2.IMWRITE_JPEG_QUALITY), 70]
                )

                if ret:
                    frame_bytes = buffer.tobytes()
                    yield (
                        b"--frame\r\n"
                        b"Content-Type: image/jpeg\r\n\r\n" + frame_bytes + b"\r\n"
                    )

                time.sleep(0.2)
                continue

            with state_lock:
                faces_to_draw = list(last_faces)

            for (x, y, w, h, face_conf, genero, gender_conf) in faces_to_draw:
                if face_conf < 0.85:
                    continue

                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)

                label = f"{genero} | rostro {face_conf:.2f} | genero {gender_conf:.2f}"
                text_y = y - 10 if y - 10 > 20 else y + h + 20

                cv2.putText(
                    frame,
                    label,
                    (x, text_y),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.55,
                    (0, 255, 0),
                    2
                )

            ret, buffer = cv2.imencode(
                ".jpg",
                frame,
                [int(cv2.IMWRITE_JPEG_QUALITY), 70]
            )

            if not ret:
                time.sleep(0.02)
                continue

            frame_bytes = buffer.tobytes()

            yield (
                b"--frame\r\n"
                b"Content-Type: image/jpeg\r\n\r\n" + frame_bytes + b"\r\n"
            )

            time.sleep(0.03)

        except GeneratorExit:
            print("Cliente de video desconectado.")
            break
        except Exception as e:
            print("Error en generate_frames:", e)
            time.sleep(0.2)


# =========================
# ESTADISTICAS PARA ANGULAR (PROTEGIDAS)
# =========================
@app.get("/registros")
def get_registros(current_user: str = Depends(get_current_user)):
    try:
        rows = read_excel_rows()
        return {
            "total": len(rows),
            "registros": rows
        }
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )


@app.get("/estadisticas")
def get_estadisticas(current_user: str = Depends(get_current_user)):
    try:
        rows = read_excel_rows()

        total_eventos = len(rows)
        total_personas = 0
        total_hombres = 0
        total_mujeres = 0

        eventos_por_dia = defaultdict(int)
        personas_por_dia = defaultdict(int)
        genero_por_dia = defaultdict(lambda: {"hombres": 0, "mujeres": 0})

        for row in rows:
            fecha_hora = row.get("FechaHora")
            rostros = row.get("RostrosDetectados") or 0
            genero_texto = row.get("Genero") or ""

            try:
                rostros = int(rostros)
            except Exception:
                rostros = 0

            fecha = "Sin fecha"
            if fecha_hora:
                fecha = str(fecha_hora).split(" ")[0]

            total_personas += rostros
            eventos_por_dia[fecha] += 1
            personas_por_dia[fecha] += rostros

            generos = [g.strip() for g in str(genero_texto).split(",") if g and str(g).strip()]

            for genero in generos:
                genero_lower = genero.lower()

                if genero_lower == "hombre":
                    total_hombres += 1
                    genero_por_dia[fecha]["hombres"] += 1
                elif genero_lower == "mujer":
                    total_mujeres += 1
                    genero_por_dia[fecha]["mujeres"] += 1

        dias_ordenados = sorted(personas_por_dia.keys())

        resumen_por_dia = []
        for dia in dias_ordenados:
            resumen_por_dia.append({
                "fecha": dia,
                "eventos": eventos_por_dia[dia],
                "personas": personas_por_dia[dia],
                "hombres": genero_por_dia[dia]["hombres"],
                "mujeres": genero_por_dia[dia]["mujeres"],
            })

        return {
            "totales": {
                "eventos": total_eventos,
                "personas": total_personas,
                "hombres": total_hombres,
                "mujeres": total_mujeres,
            },
            "por_dia": resumen_por_dia,
        }

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )


@app.get("/camera-status")
def camera_status():
    with frame_lock:
        disponible = latest_frame is not None
        segundos_sin_frames = round(time.time() - last_frame_received_time, 2) if last_frame_received_time > 0 else None

    return {
        "camera_online": disponible,
        "seconds_since_last_frame": segundos_sin_frames,
        "frame_timeout_seconds": FRAME_TIMEOUT_SECONDS,
        "camera_type": CAMERA_TYPE,
        "droidcam_url": DROIDCAM_URL,
        "camera_index": CAMERA_INDEX
    }


@app.post("/set-camera")
def set_camera(data: dict):
    global CAMERA_TYPE, CAMERA_INDEX, DROIDCAM_URL
    
    new_type = data.get("type") # "IP" o "LOCAL"
    if new_type not in ["IP", "LOCAL"]:
        raise HTTPException(status_code=400, detail="Tipo de cámara inválido. Use 'IP' o 'LOCAL'")
    
    CAMERA_TYPE = new_type
    if "index" in data and data["index"] is not None:
        CAMERA_INDEX = int(data["index"])
    if "url" in data:
        DROIDCAM_URL = data["url"]
    
    # Al cambiar el tipo, el hilo camera_reader detectará el cambio y reconectará
    return {"message": f"Cámara cambiada a {CAMERA_TYPE}", "config": {
        "type": CAMERA_TYPE,
        "index": CAMERA_INDEX,
        "url": DROIDCAM_URL
    }}


@app.get("/detecciones/activas")
def get_detecciones_activas():
    """
    Devuelve las detecciones estables actuales en memoria.
    No requiere autenticación (solo lectura del estado en RAM).
    """
    with state_lock:
        faces = list(last_faces)

    detecciones = []
    for (x, y, w, h, face_conf, genero, gender_conf) in faces:
        if face_conf < 0.85:
            continue
        if genero == "No identificado":
            continue
        detecciones.append({
            "genero": genero,
            "confianza_rostro": round(float(face_conf), 2),
            "confianza_genero": round(float(gender_conf), 2),
        })

    return {
        "timestamp": datetime.now().isoformat(),
        "detecciones": detecciones,
        "total": len(detecciones),
    }


# =========================
# EVENTOS FASTAPI
# =========================
@app.get("/video_feed")
def video_feed():
    return StreamingResponse(
        generate_frames(),
        media_type="multipart/x-mixed-replace; boundary=frame"
    )


@app.on_event("startup")
def startup_event():
    print("====================================")
    print("Iniciando API...")
    print("Carpeta actual:", os.getcwd())
    print("BASE_DIR:", BASE_DIR)
    print("MODELS_DIR:", MODELS_DIR)
    print("Ruta Excel:", EXCEL_FILE)
    print("DROIDCAM_URL:", DROIDCAM_URL)
    print("====================================")
    
    init_db()
    init_excel()
    load_models()

    thread_cam = threading.Thread(target=camera_reader, daemon=True)
    thread_cam.start()

    thread_detection = threading.Thread(target=detection_worker, daemon=True)
    thread_detection.start()

    print("Hilos iniciados correctamente.")


@app.on_event("shutdown")
def shutdown_event():
    global running
    running = False
    print("API detenida correctamente.")


# =========================
# AUTH ENDPOINTS
# =========================
@app.post("/auth/register")
def register(data: dict):
    username = data.get("username")
    email = data.get("email")
    password = data.get("password")

    if not username or not email or not password:
        raise HTTPException(status_code=400, detail="Faltan datos")

    hashed_pw = get_password_hash(password)

    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("INSERT INTO users (username, email, password) VALUES (?, ?, ?)", (username, email, hashed_pw))
        conn.commit()
        conn.close()
        return {"message": "Usuario creado correctamente"}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="El usuario o email ya existe")

@app.post("/auth/login")
def login(data: dict):
    identifier = data.get("identifier") or data.get("username")
    password   = data.get("password")

    if not identifier or not password:
        raise HTTPException(status_code=400, detail="Faltan datos")

    conn   = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    # Buscar por email o por username dependiendo del formato
    if "@" in identifier:
        cursor.execute(
            "SELECT username, password, email FROM users WHERE email = ?",
            (identifier,)
        )
    else:
        cursor.execute(
            "SELECT username, password, email FROM users WHERE username = ?",
            (identifier,)
        )

    row = cursor.fetchone()
    conn.close()

    if not row or not verify_password(password, row[1]):
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")

    username   = row[0]
    user_email = row[2]

    # Generar código 2FA
    code = f"{random.randint(100000, 999999)}"
    pending_2fa[username] = {
        "code":    code,
        "expires": time.time() + 300  # 5 minutos
    }

    # Enviar correo
    email_sent = send_2fa_email(user_email, code)

    print(f"DEBUG: Código 2FA para {username}: {code} (Email enviado: {email_sent})")

    return {
        "message":       "Código 2FA enviado al correo" if email_sent else "Error al enviar el correo (ver terminal)",
        "username":      username,
        "step":          "2fa",
        "email_preview": f"{user_email[:3]}***@{user_email.split('@')[-1]}" if user_email else None
    }


@app.post("/auth/verify-2fa")
def verify_2fa_code(data: dict):
    username = data.get("username")
    code = str(data.get("code", ""))

    if username not in pending_2fa:
        raise HTTPException(status_code=400, detail="No hay una sesión de 2FA activa")

    saved_data = pending_2fa[username]
    if time.time() > saved_data["expires"]:
        del pending_2fa[username]
        raise HTTPException(status_code=400, detail="Código expirado")

    if saved_data["code"] != code:
        raise HTTPException(status_code=400, detail="Código incorrecto")

    # Limpiar y generar tokens
    del pending_2fa[username]
    
    access_token = create_access_token(data={"sub": username})
    refresh_token = create_refresh_token(data={"sub": username})

    # Obtener el email del usuario para el perfil
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT email FROM users WHERE username = ?", (username,))
    row = cursor.fetchone()
    conn.close()
    
    email = row[0] if row else ""

    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
        "username": username,
        "email": email,
        "name": username,
        "picture": f"https://ui-avatars.com/api/?name={username}&background=1e1e2c&color=c9a96e"
    }


@app.post("/auth/google")
def google_login(data: dict):
    """Autenticación con Google OAuth. Verifica el token de Google y devuelve JWT."""
    credential = data.get("credential")
    if not credential:
        raise HTTPException(status_code=400, detail="Falta el token de Google")

    try:
        # Verificar el token de Google
        idinfo = id_token.verify_oauth2_token(
            credential,
            google_requests.Request(),
            GOOGLE_CLIENT_ID
        )

        # Extraer información del usuario de Google
        google_email = idinfo.get("email")
        google_name = idinfo.get("name", "")
        google_picture = idinfo.get("picture", "")

        if not google_email:
            raise HTTPException(status_code=400, detail="No se pudo obtener el email de Google")

        # Buscar si el usuario ya existe en la base de datos
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("SELECT username, email FROM users WHERE email = ?", (google_email,))
        row = cursor.fetchone()

        if not row:
            # Auto-registrar usuario de Google
            # Usar el nombre de Google como username (sanitizado)
            base_username = google_name.replace(" ", "_").lower() if google_name else google_email.split("@")[0]
            username = base_username

            # Verificar que el username no esté repetido
            cursor.execute("SELECT id FROM users WHERE username = ?", (username,))
            if cursor.fetchone():
                username = f"{base_username}_{random.randint(1000, 9999)}"

            # Crear contraseña aleatoria (el usuario no la necesitará porque entra con Google)
            random_password = get_password_hash(f"google_{random.randint(100000, 999999)}")

            cursor.execute(
                "INSERT INTO users (username, email, password) VALUES (?, ?, ?)",
                (username, google_email, random_password)
            )
            conn.commit()
            print(f"Usuario de Google registrado automáticamente: {username} ({google_email})")
        else:
            username = row[0]

        conn.close()

        # Generar tokens JWT directamente (sin 2FA para Google)
        access_token = create_access_token(data={"sub": username})
        refresh_token = create_refresh_token(data={"sub": username})

        return {
            "access_token": access_token,
            "refresh_token": refresh_token,
            "token_type": "bearer",
            "username": username,
            "email": google_email,
            "name": google_name,
            "picture": google_picture
        }

    except ValueError as e:
        print(f"Error verificando token de Google: {e}")
        raise HTTPException(status_code=401, detail="Token de Google inválido o expirado")
    except Exception as e:
        print(f"Error en autenticación de Google: {e}")
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")

@app.post("/auth/facebook")
def facebook_login(data: dict):
    """Autenticación con Facebook OAuth. Verifica el access token con Graph API."""
    access_token_fb = data.get("accessToken")
    if not access_token_fb:
        raise HTTPException(status_code=400, detail="Falta el token de Facebook")

    try:
        # Verificar el token con Facebook Graph API
        fb_response = requests.get(
            "https://graph.facebook.com/me",
            params={
                "fields": "id,name,email,picture.type(large)",
                "access_token": access_token_fb
            },
            timeout=10
        )

        if fb_response.status_code != 200:
            raise HTTPException(status_code=401, detail="Token de Facebook inválido")

        fb_data = fb_response.json()
        fb_email = fb_data.get("email")
        fb_name = fb_data.get("name", "")
        fb_id = fb_data.get("id", "")
        fb_picture = fb_data.get("picture", {}).get("data", {}).get("url", "")

        # Si Facebook no devuelve email, usar un email generado con el ID
        if not fb_email:
            fb_email = f"fb_{fb_id}@facebook.com"

        # Buscar si el usuario ya existe
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("SELECT username, email FROM users WHERE email = ?", (fb_email,))
        row = cursor.fetchone()

        if not row:
            # Auto-registrar usuario de Facebook
            base_username = fb_name.replace(" ", "_").lower() if fb_name else f"fb_user_{fb_id}"
            username = base_username

            # Verificar que el username no esté repetido
            cursor.execute("SELECT id FROM users WHERE username = ?", (username,))
            if cursor.fetchone():
                username = f"{base_username}_{random.randint(1000, 9999)}"

            # Contraseña aleatoria (el usuario entra con Facebook)
            random_password = get_password_hash(f"facebook_{random.randint(100000, 999999)}")

            cursor.execute(
                "INSERT INTO users (username, email, password) VALUES (?, ?, ?)",
                (username, fb_email, random_password)
            )
            conn.commit()
            print(f"Usuario de Facebook registrado automáticamente: {username} ({fb_email})")
        else:
            username = row[0]

        conn.close()

        # Generar tokens JWT directamente (sin 2FA para Facebook)
        access_token = create_access_token(data={"sub": username})
        refresh_token = create_refresh_token(data={"sub": username})

        return {
            "access_token": access_token,
            "refresh_token": refresh_token,
            "token_type": "bearer",
            "username": username,
            "email": fb_email,
            "name": fb_name,
            "picture": fb_picture
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error en autenticación de Facebook: {e}")
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")


# =========================
# ENDPOINTS PRINCIPALES
# =========================
@app.get("/")
def home():
    html_content = f"""
    <html>
        <head>
            <title>Camara en vivo</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    background: #f4f6f8;
                    padding: 20px;
                    text-align: center;
                }}
                h1 {{
                    color: #222;
                }}
                .card {{
                    max-width: 950px;
                    margin: auto;
                    background: white;
                    border-radius: 14px;
                    box-shadow: 0 4px 16px rgba(0,0,0,0.12);
                    padding: 20px;
                }}
                img {{
                    margin-top: 20px;
                    border-radius: 12px;
                    border: 2px solid #ddd;
                    max-width: 100%;
                }}
                p {{
                    margin: 8px 0;
                }}
                .links {{
                    margin-top: 18px;
                }}
                .links a {{
                    display: inline-block;
                    margin: 6px;
                    text-decoration: none;
                    color: white;
                    background: #2563eb;
                    padding: 10px 16px;
                    border-radius: 10px;
                }}
            </style>
        </head>
        <body>
            <div class="card">
                <h1>Camara con deteccion de rostro y genero</h1>
                <p><strong>Lugar configurado:</strong> {LUGAR}</p>
                <p><strong>Configuración Cámara:</strong> <span id="cam-info">{CAMERA_TYPE} ({"USB/Laptop" if CAMERA_TYPE=="LOCAL" else DROIDCAM_URL})</span></p>
                
                <div style="margin: 20px 0; padding: 15px; background: #f9f9f9; border-radius: 10px;">
                    <strong>Cambiar Cámara:</strong>
                    <button onclick="changeCam('LOCAL', 0)" style="cursor:pointer; padding: 8px; border-radius: 5px; border: 1px solid #ccc;">Laptop (Local 0)</button>
                    <button onclick="changeCam('LOCAL', 1)" style="cursor:pointer; padding: 8px; border-radius: 5px; border: 1px solid #ccc;">USB (Local 1)</button>
                    <button onclick="changeCam('IP', null)" style="cursor:pointer; padding: 8px; border-radius: 5px; border: 1px solid #ccc;">DroidCam (WiFi)</button>
                </div>

                <img src="/video" width="720" />
                
                <div class="links">
                    <a href="/detectar-rostro" target="_blank">Ver detección JSON</a>
                    <a href="/registros" target="_blank">Ver registros</a>
                    <a href="/estadisticas" target="_blank">Ver estadísticas</a>
                    <a href="/camera-status" target="_blank">Ver estado cámara</a>
                </div>
            </div>

            <script>
                async function changeCam(type, index) {{
                    const res = await fetch('/set-camera', {{
                        method: 'POST',
                        headers: {{ 'Content-Type': 'application/json' }},
                        body: JSON.stringify({{ type, index }})
                    }});
                    const data = await res.json();
                    alert(data.message);
                    location.reload();
                }}
            </script>
        </body>
    </html>
    """
    return HTMLResponse(content=html_content)


@app.get("/video")
def video():
    return StreamingResponse(
        generate_frames(),
        media_type="multipart/x-mixed-replace; boundary=frame"
    )


@app.get("/detectar-rostro")
def detectar_rostro():
    with frame_lock:
        frame = None if latest_frame is None else latest_frame.copy()

    if frame is None:
        return JSONResponse(
            status_code=500,
            content={"error": "No hay frame disponible desde DroidCam"}
        )

    faces = detectar_rostros_dnn(frame)
    detecciones = []

    for (x, y, w, h, face_conf) in faces:
        pad_x = int(w * 0.12)
        pad_y = int(h * 0.12)

        x1 = max(0, x - pad_x)
        y1 = max(0, y - pad_y)
        x2 = min(frame.shape[1], x + w + pad_x)
        y2 = min(frame.shape[0], y + h + pad_y)

        face_crop = frame[y1:y2, x1:x2]

        if face_crop is not None and face_crop.size > 0:
            genero, gender_conf = detectar_genero(face_crop)
        else:
            genero, gender_conf = "No identificado", 0.0

        if gender_conf < 0.60:
            genero = "No identificado"

        detecciones.append({
            "x": x,
            "y": y,
            "ancho": w,
            "alto": h,
            "confianza_rostro": round(face_conf, 4),
            "genero": genero,
            "confianza_genero": round(gender_conf, 4)
        })

    now = datetime.now()

    return {
        "rostros_detectados": len(detecciones),
        "hay_rostro": len(detecciones) > 0,
        "fecha_hora": now.strftime("%Y-%m-%d %H:%M:%S"),
        "lugar": LUGAR,
        "excel_file": EXCEL_FILE,
        "detecciones": detecciones
    }


@app.get("/probar-excel")
def probar_excel():
    ok = save_detection_to_excel(1, ["Hombre"])
    return {
        "ok": ok,
        "mensaje": "Fila de prueba insertada" if ok else "No se pudo insertar la fila",
        "excel_file": EXCEL_FILE
    }
    